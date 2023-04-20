####  packages

import pandas as pd
import numpy as np
import urllib.request as tr
import os
import openpyxl
from urllib.request import urlretrieve
import pydub
import webbrowser
from pydub.utils import mediainfo
from pydub import AudioSegment
import wave, array



### Obtendo o hyperlink de um worksheet do openpyxl
def try_get_hyperlink(ws, row, column): 
    try:
        return ws.cell(row=row, column=column).hyperlink.target
    except:
        return 'No Sound'


#### Lendo uma planilha do excel e retornando um pandas dataframe
def read_excel_with_hyperlink(file_name,sheet_name,column_number,column_name,rows_to_skip):
    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]
    data = ws.values
    #print(ws.cell(row=4402, column=12).hyperlink.target)
    currentRow = 1
    for eachRow in ws.iter_rows():
        ws.cell(row=currentRow, column=column_number).value = try_get_hyperlink(ws = ws, row = currentRow, column=column_number)
        currentRow += 1
    ws.cell(row=rows_to_skip, column=column_number).value ='Sounds'
    data = ws.values
    df1 = pd.DataFrame(data)
    df1 = df1.iloc[rows_to_skip-1:]
    new_header = df1.iloc[0] #grab the first row for the header
    df1 = df1[1:] #take the data less the header row
    df1.columns = new_header
    df1.reset_index(inplace=True)
    return df1

   ####### baixando dados dos lins em uma coluna com links
def download_audios(df,column_name,myPath):
    downloaded = []
    for index,r, in df.iterrows():
        if df[column_name][index] != 'No Sound':
            filename = 'audio_' + str(index) + '.wav'
            fullfilename = os.path.join(myPath, filename)
            urlretrieve(df[column_name][index], fullfilename)
            downloaded.append(str(index))
    print('Foram baixados os seguintes itens')
    print(downloaded)



##### mono to stereo

def make_stereo(file1, output):
    ifile = wave.open(file1)
    (nchannels, sampwidth, framerate, nframes, comptype, compname) = ifile.getparams()
    assert comptype == 'NONE'  # Compressed not supported yet
    array_type = {1:'B', 2: 'h', 4: 'l'}[sampwidth]
    left_channel = array.array(array_type, ifile.readframes(nframes))[::nchannels]
    ifile.close()

    stereo = 2 * left_channel
    stereo[0::2] = stereo[1::2] = left_channel

    ofile = wave.open(output, 'w')
    ofile.setparams((2, sampwidth, framerate, nframes, comptype, compname))
    ofile.writeframes(stereo.tobytes())
    ofile.close()


## convert to stereo wav

def adjust_audio(src):
    ## convert to pdf
    sound = AudioSegment.from_mp3(src)
    sound.export(src, format="wav")
    ## convert to stereo
    make_stereo(src, src)

###################


def adjust_audio_mp3(src,dest):
    ## convert to pdf
    sound = AudioSegment.from_mp3(src)
    sound.export(dest, format="wav")
    ## convert to stereo
    make_stereo(dest, dest)